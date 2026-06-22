from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "需求追踪表RTM"

# 样式定义
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 表头
headers = [
    "需求ID", "需求描述", "需求来源", "优先级", "模块", "设计文档", 
    "开发任务", "测试用例", "状态", "负责人", "备注"
]
ws.append(headers)

# 设置表头样式
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# 需求数据（从优化后的需求文档中提取）
requirements = [
    # 用户系统
    ("REQ-001", "邀请码注册制（V1.0为邀请码注册）", "产品需求说明书", "高", "用户系统", 
     "用户注册流程设计", "实现注册API、邀请码验证", "TC-001:邀请码正确可注册", "未开始", "", ""),
    ("REQ-002", "手机号/微信登录", "产品需求说明书", "高", "用户系统",
     "登录流程设计", "实现登录API、JWT Token生成", "TC-002:登录成功返回Token", "未开始", "", ""),
    ("REQ-003", "个人资料（头像、签名20字、守护灵）", "产品需求说明书", "中", "用户系统",
     "个人资料页面设计", "实现个人资料API", "TC-003:签名过AI内容安全", "未开始", "", ""),
    ("REQ-004", "账号安全（修改密码、注销账号）", "产品需求说明书", "中", "用户系统",
     "安全流程设计", "实现密码修改、账号注销API", "TC-004:密码修改成功", "未开始", "", ""),
    
    # 领地系统
    ("REQ-005", "注册时随机分配坐标（保证与邀请人在±10格内）", "产品需求说明书", "高", "领地系统",
     "坐标分配算法设计", "实现坐标分配算法、防重叠机制", "TC-005:坐标不重叠", "未开始", "", ""),
    ("REQ-006", "领地升级Lv.1~Lv.5", "产品需求说明书", "高", "领地系统",
     "升级规则设计", "实现经验值计算、升级逻辑", "TC-006:经验值正确累加", "未开始", "", ""),
    ("REQ-007", "5套固定元素图（按守护灵类型）", "产品需求说明书", "中", "领地系统",
     "元素图资源设计", "实现元素图展示逻辑", "TC-007:元素图正确展示", "未开始", "", ""),
    
    # 地图系统
    ("REQ-008", "查看坐标±10格范围内邻居", "产品需求说明书", "高", "地图系统",
     "地图展示设计", "实现邻居查询API", "TC-008:邻居列表正确", "未开始", "", ""),
    ("REQ-009", "无人区野兽潮自动防御", "产品需求说明书", "中", "地图系统",
     "野兽潮逻辑设计", "实现野兽潮定时任务、防御逻辑", "TC-009:防御成功获得金币", "未开始", "", ""),
    
    # Agent（守护灵）
    ("REQ-010", "3种类型守护灵（机械师/精灵/占星师）", "产品需求说明书", "高", "Agent系统",
     "守护灵话术设计", "实现话术模板配置表", "TC-010:话术正确触发", "未开始", "", ""),
    ("REQ-011", "固定话术模板+条件触发", "产品需求说明书", "高", "Agent系统",
     "触发机制设计", "实现触发逻辑、OpenIM自定义消息", "TC-011:Agent通知正确发送", "未开始", "", ""),
    
    # 邀请裂变
    ("REQ-012", "每个用户唯一邀请码", "产品需求说明书", "高", "邀请系统",
     "邀请码生成算法设计", "实现邀请码生成API", "TC-012:邀请码唯一", "未开始", "", ""),
    ("REQ-013", "邀请进度追踪", "产品需求说明书", "中", "邀请系统",
     "进度计算逻辑设计", "实现进度查询API", "TC-013:进度正确计数", "未开始", "", ""),
    ("REQ-014", "邀请≥3人自动生成个人小程序", "产品需求说明书", "高", "邀请系统",
     "小程序生成流程设计", "实现小程序码生成、存储逻辑", "TC-014:小程序正确生成", "未开始", "", ""),
    
    # 俱乐部（贴吧）
    ("REQ-015", "按游戏名称创建俱乐部（需Lv.2以上）", "产品需求说明书", "高", "俱乐部系统",
     "俱乐部创建流程设计", "实现创建俱乐部API、OpenIM群组创建", "TC-015:俱乐部创建成功", "未开始", "", ""),
    ("REQ-016", "发帖（500字）/回帖（200字）", "产品需求说明书", "高", "俱乐部系统",
     "帖子发布流程设计", "实现发帖、回帖API", "TC-016:帖子内容过AI审核", "未开始", "", ""),
    
    # 百货（商城）
    ("REQ-017", "虚拟道具（金币兑换：头像框200、chat气泡150）", "产品需求说明书", "中", "商城系统",
     "商城界面设计", "实现道具列表、兑换API", "TC-017:兑换成功扣除金币", "未开始", "", ""),
    ("REQ-018", "金币获取（签到、发帖、邀请、防御野兽潮）", "产品需求说明书", "高", "金币系统",
     "金币规则设计", "实现金币获取、消耗逻辑", "TC-018:金币余额正确", "未开始", "", ""),
    
    # 后台管理
    ("REQ-019", "用户管理（查看、冻结、邀请关系链）", "后台管理系统需求", "高", "后台管理系统",
     "后台管理界面设计", "实现用户管理API、前端页面", "TC-019:用户冻结成功", "未开始", "", ""),
    ("REQ-020", "内容审核（AI初审+人工复审）", "后台管理系统需求", "高", "后台管理系统",
     "审核流程设计", "实现审核API、腾讯云内容安全API对接", "TC-020:审核流程正确", "未开始", "", ""),
    
    # 非功能性需求
    ("REQ-021", "性能：API响应时间≤500ms（P95）", "产品需求说明书", "高", "非功能性需求",
     "性能测试方案设计", "实现API性能优化", "TC-021:性能测试通过", "未开始", "", ""),
    ("REQ-022", "安全性：JWT Token认证、密码bcrypt加密", "产品需求说明书", "高", "非功能性需求",
     "安全方案设计", "实现安全机制", "TC-022:安全测试通过", "未开始", "", ""),
]

# 填充数据
for req in requirements:
    ws.append(list(req))

# 设置数据样式
for row in ws.iter_rows(min_row=2, max_row=len(requirements)+1):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        # 设置状态列颜色
        if cell.column == 9:  # 状态列
            if cell.value == "未开始":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif cell.value == "进行中":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif cell.value == "已完成":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# 调整列宽
column_widths = [15, 40, 20, 10, 15, 20, 20, 25, 10, 12, 20]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 冻结首行
ws.freeze_panes = "A2"

# 添加筛选器
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

wb.save("需求追踪表RTM.xlsx")
print("需求追踪表RTM.xlsx已生成")
